"""
Module: excel_parser.py
Purpose: Extract data from Excel spreadsheets (XLSX, XLS, CSV).

Uses openpyxl for XLSX, and pandas for CSV/XLS.
Extracts sheet-level data, auto-detects tables, and
converts structured data for downstream processing.
"""

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd

from backend.models.exceptions import CorruptedFileError, DocumentParsingError
from backend.models.schemas import (
    DocumentMetadata,
    FileType,
    Page,
    ParsedDocument,
)
from backend.utils.logger import get_logger
from backend.utils.text_utils import clean_text

logger = get_logger(__name__)


class ExcelParser:
    """
    Excel/CSV parser supporting XLSX, XLS, and CSV formats.

    Extracts:
        - Sheet-level data (one "page" per sheet)
        - Auto-detected table structures
        - Cell values with type preservation
        - Spreadsheet metadata
    """

    def parse(self, file_path: str) -> ParsedDocument:
        """
        Parse a spreadsheet file.

        Args:
            file_path: Path to the spreadsheet file

        Returns:
            ParsedDocument where each sheet is a "page"

        Raises:
            FileNotFoundError: If file doesn't exist
            CorruptedFileError: If file is corrupted
            DocumentParsingError: For other parsing errors
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Spreadsheet not found: {file_path}")

        logger.info(f"Parsing spreadsheet: {path.name}")

        ext = path.suffix.lower()

        try:
            if ext == ".xlsx":
                return self._parse_xlsx(file_path)
            elif ext == ".xls":
                return self._parse_with_pandas(file_path, FileType.XLS)
            elif ext == ".csv":
                return self._parse_csv(file_path)
            else:
                raise DocumentParsingError(
                    f"Unsupported spreadsheet format: {ext}"
                )
        except (CorruptedFileError, DocumentParsingError):
            raise
        except Exception as e:
            raise DocumentParsingError(
                f"Failed to parse spreadsheet: {e}"
            ) from e

    def _parse_xlsx(self, file_path: str) -> ParsedDocument:
        """Parse XLSX using openpyxl."""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CorruptedFileError(
                f"Cannot open XLSX file: {e}"
            ) from e

        pages: List[Page] = []
        all_text_parts: List[str] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            page_num = sheet_idx + 1

            # Extract all rows
            rows: List[List[str]] = []
            for row in sheet.iter_rows(values_only=True):
                str_row = [
                    str(cell).strip() if cell is not None else ""
                    for cell in row
                ]
                rows.append(str_row)

            # Skip completely empty sheets
            if not any(any(cell for cell in row) for row in rows):
                continue

            # Remove trailing empty rows
            while rows and not any(rows[-1]):
                rows.pop()

            # Build text representation
            text_lines = [f"Sheet: {sheet_name}"]
            for row in rows:
                line = " | ".join(cell for cell in row if cell)
                if line.strip():
                    text_lines.append(line)

            text = "\n".join(text_lines)
            all_text_parts.append(text)

            # Detect headers (first non-empty row)
            headers_row_idx = 0
            for idx, row in enumerate(rows):
                if any(cell for cell in row):
                    headers_row_idx = idx
                    break

            # Build table structure
            table_data = rows[headers_row_idx:]

            pages.append(Page(
                number=page_num,
                text=text,
                tables=[table_data] if table_data else [],
                metadata={
                    "sheet_name": sheet_name,
                    "row_count": len(rows),
                    "col_count": sheet.max_column or 0,
                },
            ))

        wb.close()

        full_text = "\n\n".join(all_text_parts)

        metadata = DocumentMetadata(
            page_count=len(pages),
            file_size=Path(file_path).stat().st_size,
        )

        doc = ParsedDocument(
            source_file=file_path,
            file_type=FileType.XLSX,
            pages=pages,
            total_pages=len(pages),
            full_text=full_text,
            metadata=metadata,
        )

        logger.info(
            f"XLSX parsed: {Path(file_path).name} — "
            f"{len(pages)} sheets"
        )
        return doc

    def _parse_csv(self, file_path: str) -> ParsedDocument:
        """Parse CSV file."""
        rows: List[List[str]] = []

        try:
            # Try to detect encoding
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for row in reader:
                    str_row = [str(cell).strip() for cell in row]
                    rows.append(str_row)
        except Exception as e:
            raise CorruptedFileError(
                f"Cannot read CSV file: {e}"
            ) from e

        if not rows:
            return ParsedDocument(
                source_file=file_path,
                file_type=FileType.CSV,
                pages=[],
                total_pages=0,
                full_text="",
                metadata=DocumentMetadata(
                    file_size=Path(file_path).stat().st_size
                ),
            )

        # Build text representation
        text_lines = []
        for row in rows:
            line = " | ".join(cell for cell in row if cell)
            if line.strip():
                text_lines.append(line)

        text = "\n".join(text_lines)

        page = Page(
            number=1,
            text=text,
            tables=[rows],
            metadata={
                "row_count": len(rows),
                "col_count": max(len(r) for r in rows) if rows else 0,
            },
        )

        return ParsedDocument(
            source_file=file_path,
            file_type=FileType.CSV,
            pages=[page],
            total_pages=1,
            full_text=text,
            metadata=DocumentMetadata(
                file_size=Path(file_path).stat().st_size,
                page_count=1,
            ),
        )

    def _parse_with_pandas(
        self, file_path: str, file_type: FileType
    ) -> ParsedDocument:
        """Parse XLS using pandas (fallback for older Excel formats)."""
        try:
            dfs = pd.read_excel(file_path, sheet_name=None, dtype=str)
        except Exception as e:
            raise CorruptedFileError(
                f"Cannot read Excel file: {e}"
            ) from e

        pages: List[Page] = []
        all_text_parts: List[str] = []

        for sheet_idx, (sheet_name, df) in enumerate(dfs.items()):
            df = df.fillna("")
            page_num = sheet_idx + 1

            # Convert to list of lists
            headers = [str(col) for col in df.columns]
            rows = [headers] + df.values.tolist()
            rows = [
                [str(cell).strip() for cell in row]
                for row in rows
            ]

            # Build text
            text_lines = [f"Sheet: {sheet_name}"]
            for row in rows:
                line = " | ".join(cell for cell in row if cell)
                if line.strip():
                    text_lines.append(line)

            text = "\n".join(text_lines)
            all_text_parts.append(text)

            pages.append(Page(
                number=page_num,
                text=text,
                tables=[rows],
                metadata={"sheet_name": sheet_name},
            ))

        return ParsedDocument(
            source_file=file_path,
            file_type=file_type,
            pages=pages,
            total_pages=len(pages),
            full_text="\n\n".join(all_text_parts),
            metadata=DocumentMetadata(
                page_count=len(pages),
                file_size=Path(file_path).stat().st_size,
            ),
        )
